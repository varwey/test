# coding=utf8

import StringIO
import openpyxl
from openpyxl.cell import Cell
from contextlib import closing
from kfile.utils.kfile_logging import logger
from kfile.utils.translate.file_integrate import replace_or_origin

def extract(xlsx_content):
    wb = openpyxl.load_workbook(StringIO.StringIO(xlsx_content), use_iterators=True)
    sheet_names = wb.get_sheet_names()

    words = []
    for name in sheet_names:
        kr_logger.debug("sheet: %s"%name)
        ws = wb.get_sheet_by_name(name)
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == Cell.TYPE_STRING:
                    word = cell.internal_value
                    # cell.internal_value可能为None, openpyxl's bug???
                    if word and word.strip():
                        words.append(word.strip())

    return words


def integrate(xlsx_content, entry_list):
    """
        :param entry_list: a list store relation of src_words and their translate result, such as [(src1, trans1),(src2, trans2)]
    """
    words_dict = {}
    for entry in entry_list:
        words_dict[entry[0]] = entry[1]

    wb = openpyxl.load_workbook(StringIO.StringIO(xlsx_content))
    sheet_names = wb.get_sheet_names()
    i = 0
    for name in sheet_names:
        kr_logger.debug("sheet: %s"%name)
        # ws = wb.get_sheet_by_name(name)
        # for row in ws.iter_rows():
        ws = wb.get_sheet_by_name(name)
        for row in ws.rows:
            for cell in row:
                if cell.data_type == Cell.TYPE_STRING:
                    # cell.value可能为None, openpyxl's bug???
                    old_value = cell.value
                    if old_value and old_value.strip():
                        value_strip = old_value.strip()
                        try:
                            if value_strip:
                                cell.value, i = replace_or_origin(value_strip, i, entry_list, words_dict)
                        except Exception, e:
                            kr_logger.error("xls translate error in the integrate, lack translate result of entry(%s)" % value_strip)
                            # print "value=", cell.value ,i
                            raise Exception(e)
                        # if value_strip in word_dict and word_dict[value_strip]:
                        #     cell.value = word_dict[value_strip]

    with closing(StringIO.StringIO()) as sio:
        wb.save(sio)
        return sio.getvalue()



if __name__ == "__main__":
    # from contextlib import closing
    # from kfile.utils.fileUtil import read_file, write_file
    #
    # filepath = u"E:\\xls\\新手七天乐.xlsx"
    # destpath = u"E:\\xls\\result.xlsx"
    #
    # content = read_file(filepath)
    # words = extract(content)
    # print len(words)
    #
    # word_dict = dict((word, word + u"translated") for word in words)
    # write_file(integrate(content, word_dict), destpath)
    import os
    from kfile.utils.fileUtil import read_file, write_file

    # dirname = "E:\\xls"
    # excel_file_path = os.path.join(dirname, "AOE-活动库.xls").decode("utf8").encode("gbk")
    excel_file_path = "/root/share/07excel.xlsx"
    print type(excel_file_path)
    print os.path.exists(excel_file_path)
    content = read_file(excel_file_path)

    print len(content)

    word_list = extract(content)
    i = 0
    word_list2 = extract(content)
    print len(word_list)
    for i in range(len(word_list)):
        if word_list[i] != word_list2[i]:
            print "aaaaaaaaaaaaaaaaaaaaaaa", word_list[i], word_list2[i]
        if i <20 :
            print word_list[i]
    print "ddddddddaaaaaaa", len(word_list)
    # word_dict = dict()
    # for word in word_list:
    #     word_dict[word] = word + "translated"
    #
    # print type(content)
    w_content = integrate(content, word_list)
    # print len(w_content)
    # write_file(w_content, os.path.join(dirname, "result.xls"))